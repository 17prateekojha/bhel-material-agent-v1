
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, Reference


# =========================================================
# CONSTANTS
# =========================================================

TITLE = "BHEL TBG HVDC Nagpur"
SUBTITLE = "Material Management Status Report"

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

TITLE_FONT = Font(
    size=16,
    bold=True,
)

SUBTITLE_FONT = Font(
    size=11,
    italic=True,
)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2"),
)


# =========================================================
# GENERIC HELPERS
# =========================================================

def _safe_value(obj, attribute: str, default=""):
    """Safely read an attribute from a SQLAlchemy model/object."""
    return getattr(obj, attribute, default)


def _write_title(ws, title: str = TITLE, subtitle: str = SUBTITLE):
    """Write standard BHEL report title."""
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT

    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT

    ws["A3"] = (
        "Report Generated: "
        + datetime.now().strftime("%d-%b-%Y %H:%M:%S")
    )

    ws["A1"].alignment = Alignment(horizontal="left")
    ws["A2"].alignment = Alignment(horizontal="left")
    ws["A3"].alignment = Alignment(horizontal="left")


def _write_headers(ws, row_number: int, headers: list[str]):
    """Write and format a header row."""
    for column_number, header in enumerate(headers, start=1):
        cell = ws.cell(
            row=row_number,
            column=column_number,
            value=header,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = THIN_BORDER


def _format_data_range(
    ws,
    start_row: int,
    end_row: int,
    start_column: int,
    end_column: int,
):
    """Apply borders/alignment to a data range."""
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_column,
        max_col=end_column,
    ):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )


def _auto_width(ws):
    """Automatically size worksheet columns."""
    for column_cells in ws.columns:
        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:
            try:
                value_length = len(str(cell.value or ""))
                max_length = max(
                    max_length,
                    value_length,
                )
            except Exception:
                pass

        ws.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            40,
        )


def _finish_sheet(
    ws,
    header_row: int,
    freeze_cell: str = "A5",
):
    """Apply common worksheet settings."""
    ws.freeze_panes = freeze_cell

    ws.auto_filter.ref = (
        f"A{header_row}:"
        f"{get_column_letter(ws.max_column)}{ws.max_row}"
    )

    ws.sheet_view.showGridLines = False

    _auto_width(ws)


# =========================================================
# MATERIAL MASTER
# =========================================================

def add_material_master_sheet(
    wb: Workbook,
    materials: Iterable,
):
    ws = wb.create_sheet("Material Master")

    _write_title(ws)

    headers = [
        "Material ID",
        "Material Description",
        "Category",
        "Unit",
        "Initial Quantity",
        "Current Quantity",
        "Supplier",
        "PO Number",
        "GRN Number",
        "Status",
        "Last Updated",
    ]

    header_row = 5
    _write_headers(ws, header_row, headers)

    row = header_row + 1

    for material in materials:
        ws.cell(row, 1, _safe_value(material, "material_id"))
        ws.cell(row, 2, _safe_value(material, "description"))
        ws.cell(row, 3, _safe_value(material, "category"))
        ws.cell(row, 4, _safe_value(material, "unit"))
        ws.cell(row, 5, _safe_value(material, "quantity", 0))
        ws.cell(row, 6, _safe_value(material, "quantity", 0))
        ws.cell(row, 7, _safe_value(material, "supplier"))
        ws.cell(row, 8, _safe_value(material, "po_number"))
        ws.cell(row, 9, _safe_value(material, "grn_number"))
        ws.cell(row, 10, "ACTIVE")
        ws.cell(
            row,
            11,
            _safe_value(
                material,
                "updated_at",
                datetime.now(),
            ),
        )

        row += 1

    if row > header_row + 1:
        _format_data_range(
            ws,
            header_row + 1,
            row - 1,
            1,
            len(headers),
        )

    _finish_sheet(ws, header_row)

    return ws


# =========================================================
# MATERIAL RECEIPT
# =========================================================

def add_material_receipt_sheet(
    wb: Workbook,
    receipts: Iterable | None = None,
):
    ws = wb.create_sheet("Material Receipt")

    _write_title(ws)

    headers = [
        "Receipt ID",
        "Receipt Date",
        "Material ID",
        "Material Description",
        "Supplier",
        "PO Number",
        "GRN Number",
        "Received Quantity",
        "Unit",
        "Location",
        "Received By",
        "Remarks",
    ]

    header_row = 5
    _write_headers(ws, header_row, headers)

    row = header_row + 1

    for receipt in receipts or []:
        values = [
            _safe_value(receipt, "receipt_id"),
            _safe_value(receipt, "receipt_date"),
            _safe_value(receipt, "material_id"),
            _safe_value(receipt, "description"),
            _safe_value(receipt, "supplier"),
            _safe_value(receipt, "po_number"),
            _safe_value(receipt, "grn_number"),
            _safe_value(receipt, "quantity", 0),
            _safe_value(receipt, "unit"),
            _safe_value(receipt, "location"),
            _safe_value(receipt, "received_by"),
            _safe_value(receipt, "remarks"),
        ]

        for column, value in enumerate(values, start=1):
            ws.cell(row, column, value)

        row += 1

    if row > header_row + 1:
        _format_data_range(
            ws,
            header_row + 1,
            row - 1,
            1,
            len(headers),
        )

    _finish_sheet(ws, header_row)

    return ws


# =========================================================
# STORE BALANCE
# =========================================================

def add_store_balance_sheet(
    wb: Workbook,
    materials: Iterable,
):
    ws = wb.create_sheet("Store Balance")

    _write_title(ws)

    headers = [
        "Material ID",
        "Material Description",
        "Unit",
        "Total Received",
        "Total Issued",
        "Total Transferred",
        "Adjustment",
        "Current Balance",
    ]

    header_row = 5
    _write_headers(ws, header_row, headers)

    row = header_row + 1

    for material in materials:
        ws.cell(
            row,
            1,
            _safe_value(material, "material_id"),
        )

        ws.cell(
            row,
            2,
            _safe_value(material, "description"),
        )

        ws.cell(
            row,
            3,
            _safe_value(material, "unit"),
        )

        ws.cell(
            row,
            4,
            _safe_value(material, "total_received", 0),
        )

        ws.cell(
            row,
            5,
            _safe_value(material, "total_issued", 0),
        )

        ws.cell(
            row,
            6,
            _safe_value(material, "total_transferred", 0),
        )

        ws.cell(
            row,
            7,
            _safe_value(material, "adjustment", 0),
        )

        # Excel formula
        ws.cell(
            row,
            8,
            f"=D{row}-E{row}-F{row}+G{row}",
        )

        row += 1

    if row > header_row + 1:
        _format_data_range(
            ws,
            header_row + 1,
            row - 1,
            1,
            len(headers),
        )

    _finish_sheet(ws, header_row)

    return ws


# =========================================================
# TRANSACTIONS
# =========================================================

def add_transactions_sheet(
    wb: Workbook,
    transactions: Iterable | None = None,
):
    ws = wb.create_sheet("Transactions")

    _write_title(ws)

    headers = [
        "Transaction ID",
        "Transaction Date",
        "Material ID",
        "Transaction Type",
        "Quantity",
        "Unit",
        "From Location",
        "To Location",
        "Reference",
        "User",
        "Remarks",
    ]

    header_row = 5
    _write_headers(ws, header_row, headers)

    row = header_row + 1

    for transaction in transactions or []:
        values = [
            _safe_value(transaction, "transaction_id"),
            _safe_value(transaction, "transaction_date"),
            _safe_value(transaction, "material_id"),
            _safe_value(transaction, "transaction_type"),
            _safe_value(transaction, "quantity", 0),
            _safe_value(transaction, "unit"),
            _safe_value(transaction, "from_location"),
            _safe_value(transaction, "to_location"),
            _safe_value(transaction, "reference"),
            _safe_value(transaction, "user"),
            _safe_value(transaction, "remarks"),
        ]

        for column, value in enumerate(values, start=1):
            ws.cell(row, column, value)

        row += 1

    # Transaction type drop-down
    validation = DataValidation(
        type="list",
        formula1='"Receipt,Issue,Handover,Transfer,Adjustment-In,Adjustment-Out"',
        allow_blank=True,
    )

    ws.add_data_validation(validation)

    validation.add(
        f"D{header_row + 1}:D{max(row - 1, header_row + 1000)}"
    )

    if row > header_row + 1:
        _format_data_range(
            ws,
            header_row + 1,
            row - 1,
            1,
            len(headers),
        )

    _finish_sheet(ws, header_row)

    return ws


# =========================================================
# RECONCILIATION
# =========================================================

def add_reconciliation_sheet(
    wb: Workbook,
    reconciliations: Iterable | None = None,
):
    ws = wb.create_sheet("Reconciliation")

    _write_title(ws)

    headers = [
        "Material ID",
        "Material Description",
        "System Quantity",
        "Physical Quantity",
        "Difference",
        "Tolerance",
        "Exception",
        "Reconciliation Date",
        "Verified By",
        "Remarks",
    ]

    header_row = 5
    _write_headers(ws, header_row, headers)

    row = header_row + 1

    for item in reconciliations or []:
        ws.cell(
            row,
            1,
            _safe_value(item, "material_id"),
        )

        ws.cell(
            row,
            2,
            _safe_value(item, "description"),
        )

        ws.cell(
            row,
            3,
            _safe_value(item, "system_quantity", 0),
        )

        ws.cell(
            row,
            4,
            _safe_value(item, "physical_quantity", 0),
        )

        # Difference formula
        ws.cell(
            row,
            5,
            f"=D{row}-C{row}",
        )

        ws.cell(
            row,
            6,
            _safe_value(item, "tolerance", 0),
        )

        # Exception formula
        ws.cell(
            row,
            7,
            (
                f'=IF(ABS(E{row})<=F{row},'
                f'"OK",'
                f'IF(E{row}<0,"SHORTAGE","EXCESS"))'
            ),
        )

        ws.cell(
            row,
            8,
            _safe_value(item, "reconciliation_date"),
        )

        ws.cell(
            row,
            9,
            _safe_value(item, "verified_by"),
        )

        ws.cell(
            row,
            10,
            _safe_value(item, "remarks"),
        )

        row += 1

    if row > header_row + 1:
        _format_data_range(
            ws,
            header_row + 1,
            row - 1,
            1,
            len(headers),
        )

    _finish_sheet(ws, header_row)

    return ws


# =========================================================
# SUPPLIER / PO
# =========================================================

def add_supplier_po_sheet(
    wb: Workbook,
    supplier_pos: Iterable | None = None,
):
    ws = wb.create_sheet("Supplier-PO")

    _write_title(ws)

    headers = [
        "Supplier ID",
        "Supplier Name",
        "PO Number",
        "PO Date",
        "Material ID",
        "Material Description",
        "Ordered Quantity",
        "Received Quantity",
        "Balance PO Quantity",
        "GRN Number",
        "PO Status",
    ]

    header_row = 5
    _write_headers(ws, header_row, headers)

    row = header_row + 1

    for item in supplier_pos or []:
        ws.cell(
            row,
            1,
            _safe_value(item, "supplier_id"),
        )

        ws.cell(
            row,
            2,
            _safe_value(item, "supplier_name"),
        )

        ws.cell(
            row,
            3,
            _safe_value(item, "po_number"),
        )

        ws.cell(
            row,
            4,
            _safe_value(item, "po_date"),
        )

        ws.cell(
            row,
            5,
            _safe_value(item, "material_id"),
        )

        ws.cell(
            row,
            6,
            _safe_value(item, "description"),
        )

        ws.cell(
            row,
            7,
            _safe_value(item, "ordered_quantity", 0),
        )

        ws.cell(
            row,
            8,
            _safe_value(item, "received_quantity", 0),
        )

        # PO balance formula
        ws.cell(
            row,
            9,
            f"=G{row}-H{row}",
        )

        ws.cell(
            row,
            10,
            _safe_value(item, "grn_number"),
        )

        ws.cell(
            row,
            11,
            _safe_value(item, "po_status"),
        )

        row += 1

    if row > header_row + 1:
        _format_data_range(
            ws,
            header_row + 1,
            row - 1,
            1,
            len(headers),
        )

    _finish_sheet(ws, header_row)

    return ws


# =========================================================
# DASHBOARD
# =========================================================

def add_dashboard_sheet(
    wb: Workbook,
    materials: Iterable,
    transactions: Iterable | None = None,
    reconciliations: Iterable | None = None,
):
    ws = wb.create_sheet("Dashboard", 0)

    _write_title(
        ws,
        TITLE,
        "Material Management Dashboard",
    )

    materials_list = list(materials)
    transactions_list = list(transactions or [])
    reconciliation_list = list(reconciliations or [])

    total_materials = len(materials_list)

    total_received = sum(
        _safe_value(
            material,
            "total_received",
            0,
        )
        or 0
        for material in materials_list
    )

    total_issued = sum(
        _safe_value(
            material,
            "total_issued",
            0,
        )
        or 0
        for material in materials_list
    )

    total_balance = sum(
        _safe_value(
            material,
            "quantity",
            0,
        )
        or 0
        for material in materials_list
    )

    exception_count = len(
        [
            item
            for item in reconciliation_list
            if _safe_value(item, "exception")
            not in ("", None, "OK")
        ]
    )

    metrics = [
        ("Total Materials", total_materials),
        ("Total Received", total_received),
        ("Total Issued", total_issued),
        ("Current Stock", total_balance),
        ("Transactions", len(transactions_list)),
        ("Reconciliation Exceptions", exception_count),
    ]

    start_row = 5

    for index, (label, value) in enumerate(metrics):
        column = 1 + (index % 3) * 3
        row = start_row + (index // 3) * 4

        label_cell = ws.cell(
            row,
            column,
            label,
        )

        value_cell = ws.cell(
            row + 1,
            column,
            value,
        )

        label_cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        label_cell.fill = HEADER_FILL
        label_cell.alignment = Alignment(
            horizontal="center"
        )

        value_cell.font = Font(
            bold=True,
            size=18,
        )

        value_cell.alignment = Alignment(
            horizontal="center"
        )

    ws["A14"] = "Report Information"
    ws["A14"].font = Font(bold=True)

    ws["A15"] = "Generated"
    ws["B15"] = datetime.now().strftime(
        "%d-%b-%Y %H:%M:%S"
    )

    ws["A16"] = "Project"
    ws["B16"] = "BHEL TBG HVDC Nagpur"

    ws["A17"] = "Report"
    ws["B17"] = "Material Management Status"

    ws.sheet_view.showGridLines = False

    for column in range(1, 10):
        ws.column_dimensions[
            get_column_letter(column)
        ].width = 18

    return ws


# =========================================================
# REFERENCE / LISTS
# =========================================================

def add_lists_sheet(wb: Workbook):
    ws = wb.create_sheet("Lists")

    ws["A1"] = "Transaction Types"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL

    transaction_types = [
        "Receipt",
        "Issue",
        "Handover",
        "Transfer",
        "Adjustment-In",
        "Adjustment-Out",
    ]

    for index, value in enumerate(
        transaction_types,
        start=2,
    ):
        ws.cell(index, 1, value)

    ws["C1"] = "Status"
    ws["C1"].font = HEADER_FONT
    ws["C1"].fill = HEADER_FILL

    statuses = [
        "ACTIVE",
        "CLOSED",
        "HOLD",
        "PENDING",
    ]

    for index, value in enumerate(
        statuses,
        start=2,
    ):
        ws.cell(index, 3, value)

    ws.sheet_state = "hidden"

    return ws


# =========================================================
# MAIN REPORT GENERATOR
# =========================================================

def generate_material_status_report(
    materials: Iterable,
    receipts: Iterable | None = None,
    transactions: Iterable | None = None,
    reconciliations: Iterable | None = None,
    supplier_pos: Iterable | None = None,
) -> BytesIO:
    """
    Generate the complete BHEL Material Management Excel report.

    Returns:
        BytesIO containing the generated .xlsx workbook.
    """

    wb = Workbook()

    # Remove default sheet.
    default_sheet = wb.active
    wb.remove(default_sheet)

    materials = list(materials)
    receipts = list(receipts or [])
    transactions = list(transactions or [])
    reconciliations = list(reconciliations or [])
    supplier_pos = list(supplier_pos or [])

    # Dashboard
    add_dashboard_sheet(
        wb,
        materials,
        transactions,
        reconciliations,
    )

    # Main sheets
    add_material_master_sheet(
        wb,
        materials,
    )

    add_material_receipt_sheet(
        wb,
        receipts,
    )

    add_store_balance_sheet(
        wb,
        materials,
    )

    add_transactions_sheet(
        wb,
        transactions,
    )

    add_reconciliation_sheet(
        wb,
        reconciliations,
    )

    add_supplier_po_sheet(
        wb,
        supplier_pos,
    )

    # Hidden reference sheet
    add_lists_sheet(wb)

    # Save into memory
    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output

